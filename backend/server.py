from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from bson import ObjectId
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import urllib.request

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
SECRET_KEY = os.environ.get('JWT_SECRET', 'ibn-khaldoun-secret-key-2026')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24

# School Logo URL
SCHOOL_LOGO_URL = "https://customer-assets.emergentagent.com/job_7beeca3e-b314-4460-83b0-e8b48115e3c8/artifacts/80ve4mnx_1000219663.jpg"

# Security
security = HTTPBearer()

# Create the main app
app = FastAPI(title="نظام ثانوية ابن خلدون الخاصة")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ==================== MODELS ====================

class UserBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    username: str
    full_name: str
    role: str  # admin, supervisor, teacher, student
    gender: Optional[str] = "male"

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class UserResponse(UserBase):
    id: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class PasswordChange(BaseModel):
    current_password: str
    new_password: str

# Student Models
class ParentInfo(BaseModel):
    father_name: str
    mother_name: str
    father_phone: str
    mother_phone: str

class StudentBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    full_name: str
    gender: str
    birth_date: str
    class_name: str
    section: str
    address: str
    parent_info: ParentInfo
    registration_date: Optional[str] = None

class StudentCreate(StudentBase):
    pass

class StudentResponse(StudentBase):
    id: str
    user_id: Optional[str] = None
    username: Optional[str] = None

# Teacher Models
class TeacherBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    full_name: str
    gender: str
    subject: str
    phone: str
    hourly_rate: float
    total_hours: float = 0
    bonus: float = 0
    deductions: float = 0
    assigned_classes: List[dict] = []  # [{class_name, section}]

class TeacherCreate(TeacherBase):
    pass

class TeacherResponse(TeacherBase):
    id: str
    user_id: Optional[str] = None
    username: Optional[str] = None
    calculated_salary: Optional[float] = None

# Grade Models - New structure
class GradeEntry(BaseModel):
    subject: str
    oral_skills: float = 0  # مهارات شفهية
    homework: float = 0  # وظائف
    quiz: float = 0  # مذاكرة
    exam: float = 0  # امتحان
    total: float = 0  # المجموع

class GradeCreate(BaseModel):
    student_id: str
    subject: str
    oral_skills: float = 0
    homework: float = 0
    quiz: float = 0
    exam: float = 0
    semester: str
    academic_year: str

# Schedule Models
class ScheduleItem(BaseModel):
    day: str
    period: int
    subject: str
    teacher_name: str
    class_name: str
    section: str

class ExamScheduleItem(BaseModel):
    date: str
    time: str
    subject: str
    class_name: str
    section: str
    duration: str

# Class Fee Model
class ClassFee(BaseModel):
    class_name: str
    annual_fee: float

# Attendance Models
class AttendanceCreate(BaseModel):
    student_id: str
    date: str
    status: str
    notes: Optional[str] = None

# Teacher Attendance
class TeacherAttendanceCreate(BaseModel):
    teacher_id: str
    date: str
    status: str
    notes: Optional[str] = None

# Announcement Models
class AnnouncementCreate(BaseModel):
    title: str
    content: str
    target_audience: str
    target_gender: Optional[str] = None  # male, female, all

class AnnouncementResponse(BaseModel):
    id: str
    title: str
    content: str
    target_audience: str
    target_gender: Optional[str] = None
    is_active: bool
    created_at: str
    created_by: str

# Financial Models
class FinancialCreate(BaseModel):
    student_id: str
    total_fee: float
    discount: float = 0

class PaymentCreate(BaseModel):
    student_id: str
    amount: float
    date: str
    notes: Optional[str] = None

# Settings Model
class SchoolSettings(BaseModel):
    school_name: str = "ثانوية ابن خلدون الخاصة"
    whatsapp_number: str = "0964803354"
    address: str = "حمص - سوريا"
    phone: str = ""
    email: str = ""

# ==================== HELPER FUNCTIONS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def generate_simple_username(role: str, name: str, counter: int = 0) -> str:
    """Generate simple readable username"""
    # Take first name only
    first_name = name.split()[0] if name else "user"
    if counter > 0:
        return f"{first_name}{counter}"
    return first_name

async def get_next_username(role: str, name: str) -> str:
    """Get next available simple username"""
    base_name = name.split()[0] if name else "user"
    counter = 0
    while True:
        username = generate_simple_username(role, name, counter)
        existing = await db.users.find_one({"username": username})
        if not existing:
            return username
        counter += 1

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def require_admin_or_supervisor(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "supervisor"]:
        raise HTTPException(status_code=403, detail="Admin or Supervisor access required")
    return current_user

async def require_teacher_or_above(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in ["admin", "supervisor", "teacher"]:
        raise HTTPException(status_code=403, detail="Teacher or higher access required")
    return current_user

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="اسم المستخدم غير موجود")
    if not verify_password(credentials.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="كلمة المرور غير صحيحة")
    
    access_token = create_access_token({"sub": user["id"], "role": user["role"]})
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(
            id=user["id"],
            username=user["username"],
            full_name=user["full_name"],
            role=user["role"],
            gender=user.get("gender", "male")
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        username=current_user["username"],
        full_name=current_user["full_name"],
        role=current_user["role"],
        gender=current_user.get("gender", "male")
    )

@api_router.post("/auth/change-password")
async def change_password(data: PasswordChange, current_user: dict = Depends(get_current_user)):
    if not verify_password(data.current_password, current_user.get("password_hash", "")):
        raise HTTPException(status_code=400, detail="كلمة المرور الحالية غير صحيحة")
    
    new_hash = hash_password(data.new_password)
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"password_hash": new_hash}})
    return {"message": "تم تغيير كلمة المرور بنجاح"}

# ==================== STUDENT ROUTES ====================

@api_router.get("/students")
async def get_students(gender: Optional[str] = None, class_name: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "student":
        student = await db.students.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if student:
            user = await db.users.find_one({"id": student.get("user_id")}, {"_id": 0})
            student["username"] = user.get("username") if user else None
            return [student]
        return []
    
    query = {}
    
    # Filter by gender for teachers (they see only their assigned gender)
    if current_user["role"] == "teacher":
        teacher = await db.teachers.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if teacher and teacher.get("assigned_classes"):
            # Get classes this teacher is assigned to
            assigned = teacher.get("assigned_classes", [])
            class_sections = [(a["class_name"], a["section"]) for a in assigned]
            query["$or"] = [{"class_name": c, "section": s} for c, s in class_sections]
    
    if gender:
        query["gender"] = gender
    if class_name:
        query["class_name"] = class_name
    
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    for student in students:
        user = await db.users.find_one({"id": student.get("user_id")}, {"_id": 0})
        student["username"] = user.get("username") if user else None
    return students

@api_router.get("/students/{student_id}")
async def get_student(student_id: str, current_user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    user = await db.users.find_one({"id": student.get("user_id")}, {"_id": 0})
    student["username"] = user.get("username") if user else None
    return student

@api_router.get("/students/{student_id}/details")
async def get_student_details(student_id: str, current_user: dict = Depends(get_current_user)):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    user = await db.users.find_one({"id": student.get("user_id")}, {"_id": 0})
    student["username"] = user.get("username") if user else None
    
    grades = await db.grades.find({"student_id": student_id}, {"_id": 0}).to_list(100)
    attendance = await db.attendance.find({"student_id": student_id}, {"_id": 0}).to_list(1000)
    
    total_days = len(attendance)
    present_days = len([a for a in attendance if a["status"] == "present"])
    absent_days = len([a for a in attendance if a["status"] == "absent"])
    late_days = len([a for a in attendance if a["status"] == "late"])
    
    financial = await db.financials.find_one({"student_id": student_id}, {"_id": 0})
    if financial:
        total_paid = sum(p.get("amount", 0) for p in financial.get("payments", []))
        financial["total_paid"] = total_paid
        financial["remaining"] = financial.get("total_fee", 0) - financial.get("discount", 0) - total_paid
    
    return {
        "student": student,
        "grades": grades,
        "attendance": {
            "records": attendance,
            "stats": {
                "total_days": total_days,
                "present_days": present_days,
                "absent_days": absent_days,
                "late_days": late_days,
                "attendance_rate": round((present_days / total_days * 100) if total_days > 0 else 0, 1)
            }
        },
        "financial": financial
    }

@api_router.post("/students")
async def create_student(student: StudentCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    student_dict = student.model_dump()
    student_dict["id"] = str(uuid.uuid4())
    student_dict["registration_date"] = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    user_id = str(uuid.uuid4())
    username = await get_next_username("student", student.full_name)
    
    user_doc = {
        "id": user_id,
        "username": username,
        "password_hash": hash_password("123456"),
        "full_name": student.full_name,
        "role": "student",
        "gender": student.gender
    }
    await db.users.insert_one(user_doc)
    student_dict["user_id"] = user_id
    
    await db.students.insert_one(student_dict)
    
    # Auto-create financial record with class fee
    class_fee = await db.class_fees.find_one({"class_name": student.class_name}, {"_id": 0})
    fee_amount = class_fee.get("annual_fee", 0) if class_fee else 0
    
    await db.financials.insert_one({
        "id": str(uuid.uuid4()),
        "student_id": student_dict["id"],
        "student_name": student.full_name,
        "class_name": student.class_name,
        "total_fee": fee_amount,
        "discount": 0,
        "payments": []
    })
    
    student_dict["username"] = username
    return student_dict

@api_router.put("/students/{student_id}")
async def update_student(student_id: str, student: StudentCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    existing = await db.students.find_one({"id": student_id})
    if not existing:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    update_data = student.model_dump()
    await db.students.update_one({"id": student_id}, {"$set": update_data})
    
    if existing.get("user_id"):
        await db.users.update_one(
            {"id": existing["user_id"]},
            {"$set": {"full_name": student.full_name, "gender": student.gender}}
        )
    
    updated = await db.students.find_one({"id": student_id}, {"_id": 0})
    user = await db.users.find_one({"id": updated.get("user_id")}, {"_id": 0})
    updated["username"] = user.get("username") if user else None
    return updated

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str, current_user: dict = Depends(require_admin_or_supervisor)):
    student = await db.students.find_one({"id": student_id})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    if student.get("user_id"):
        await db.users.delete_one({"id": student["user_id"]})
    
    await db.grades.delete_many({"student_id": student_id})
    await db.attendance.delete_many({"student_id": student_id})
    await db.financials.delete_one({"student_id": student_id})
    await db.students.delete_one({"id": student_id})
    return {"message": "تم حذف الطالب بنجاح"}

# ==================== HONOR ROLL (لوحة الشرف) ====================

@api_router.get("/honor-roll")
async def get_honor_roll(class_name: Optional[str] = None, gender: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get top 3 students per class"""
    query = {}
    if class_name:
        query["class_name"] = class_name
    if gender:
        query["gender"] = gender
    
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    
    # Get grades for all students and calculate totals
    student_totals = []
    for student in students:
        grades = await db.grades.find({"student_id": student["id"], "is_published": True}, {"_id": 0}).to_list(100)
        total = 0
        subjects_count = 0
        for grade_doc in grades:
            for g in grade_doc.get("grades", []):
                total += g.get("total", 0)
                subjects_count += 1
        
        average = total / subjects_count if subjects_count > 0 else 0
        student_totals.append({
            "student": student,
            "total": total,
            "average": round(average, 2),
            "subjects_count": subjects_count
        })
    
    # Group by class and get top 3
    from collections import defaultdict
    by_class = defaultdict(list)
    for st in student_totals:
        by_class[st["student"]["class_name"]].append(st)
    
    honor_roll = {}
    for cls, students_list in by_class.items():
        sorted_students = sorted(students_list, key=lambda x: x["average"], reverse=True)[:3]
        honor_roll[cls] = sorted_students
    
    return honor_roll

# ==================== TEACHER ROUTES ====================

@api_router.get("/teachers")
async def get_teachers(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "teacher":
        teacher = await db.teachers.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if teacher:
            teacher["calculated_salary"] = (teacher["hourly_rate"] * teacher["total_hours"]) + teacher["bonus"] - teacher["deductions"]
            user = await db.users.find_one({"id": teacher.get("user_id")}, {"_id": 0})
            teacher["username"] = user.get("username") if user else None
            return [teacher]
        return []
    
    teachers = await db.teachers.find({}, {"_id": 0}).to_list(1000)
    result = []
    for t in teachers:
        t["calculated_salary"] = (t["hourly_rate"] * t["total_hours"]) + t["bonus"] - t["deductions"]
        user = await db.users.find_one({"id": t.get("user_id")}, {"_id": 0})
        t["username"] = user.get("username") if user else None
        result.append(t)
    return result

@api_router.post("/teachers")
async def create_teacher(teacher: TeacherCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    teacher_dict = teacher.model_dump()
    teacher_dict["id"] = str(uuid.uuid4())
    
    user_id = str(uuid.uuid4())
    username = await get_next_username("teacher", teacher.full_name)
    
    user_doc = {
        "id": user_id,
        "username": username,
        "password_hash": hash_password("123456"),
        "full_name": teacher.full_name,
        "role": "teacher",
        "gender": teacher.gender
    }
    await db.users.insert_one(user_doc)
    teacher_dict["user_id"] = user_id
    
    await db.teachers.insert_one(teacher_dict)
    teacher_dict["calculated_salary"] = (teacher_dict["hourly_rate"] * teacher_dict["total_hours"]) + teacher_dict["bonus"] - teacher_dict["deductions"]
    teacher_dict["username"] = username
    return teacher_dict

@api_router.put("/teachers/{teacher_id}")
async def update_teacher(teacher_id: str, teacher: TeacherCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    existing = await db.teachers.find_one({"id": teacher_id})
    if not existing:
        raise HTTPException(status_code=404, detail="الأستاذ غير موجود")
    
    update_data = teacher.model_dump()
    await db.teachers.update_one({"id": teacher_id}, {"$set": update_data})
    
    if existing.get("user_id"):
        await db.users.update_one(
            {"id": existing["user_id"]},
            {"$set": {"full_name": teacher.full_name, "gender": teacher.gender}}
        )
    
    updated = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    updated["calculated_salary"] = (updated["hourly_rate"] * updated["total_hours"]) + updated["bonus"] - updated["deductions"]
    user = await db.users.find_one({"id": updated.get("user_id")}, {"_id": 0})
    updated["username"] = user.get("username") if user else None
    return updated

@api_router.delete("/teachers/{teacher_id}")
async def delete_teacher(teacher_id: str, current_user: dict = Depends(require_admin_or_supervisor)):
    teacher = await db.teachers.find_one({"id": teacher_id})
    if not teacher:
        raise HTTPException(status_code=404, detail="الأستاذ غير موجود")
    
    if teacher.get("user_id"):
        await db.users.delete_one({"id": teacher["user_id"]})
    
    await db.teachers.delete_one({"id": teacher_id})
    return {"message": "تم حذف الأستاذ بنجاح"}

@api_router.put("/teachers/{teacher_id}/assign-classes")
async def assign_teacher_classes(teacher_id: str, data: dict, current_user: dict = Depends(require_admin_or_supervisor)):
    """Assign classes/sections to a teacher (Supervisor function)"""
    teacher = await db.teachers.find_one({"id": teacher_id})
    if not teacher:
        raise HTTPException(status_code=404, detail="الأستاذ غير موجود")
    
    assigned_classes = data.get("assigned_classes", [])
    await db.teachers.update_one({"id": teacher_id}, {"$set": {"assigned_classes": assigned_classes}})
    return {"message": "تم تحديث الشعب المسؤول عنها"}

# ==================== TEACHER ATTENDANCE (حضور المعلمين) ====================

@api_router.get("/teacher-attendance")
async def get_teacher_attendance(date: Optional[str] = None, current_user: dict = Depends(require_admin_or_supervisor)):
    query = {}
    if date:
        query["date"] = date
    records = await db.teacher_attendance.find(query, {"_id": 0}).to_list(1000)
    return records

@api_router.post("/teacher-attendance")
async def mark_teacher_attendance(record: TeacherAttendanceCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    teacher = await db.teachers.find_one({"id": record.teacher_id}, {"_id": 0})
    if not teacher:
        raise HTTPException(status_code=404, detail="المعلم غير موجود")
    
    existing = await db.teacher_attendance.find_one({
        "teacher_id": record.teacher_id,
        "date": record.date
    })
    
    if existing:
        await db.teacher_attendance.update_one(
            {"teacher_id": record.teacher_id, "date": record.date},
            {"$set": {"status": record.status, "notes": record.notes}}
        )
    else:
        await db.teacher_attendance.insert_one({
            "id": str(uuid.uuid4()),
            "teacher_id": record.teacher_id,
            "teacher_name": teacher["full_name"],
            "date": record.date,
            "status": record.status,
            "notes": record.notes
        })
    
    return {"message": "تم تسجيل حضور المعلم"}

# ==================== SUPERVISOR ROUTES ====================

@api_router.post("/supervisors")
async def create_supervisor(data: dict, current_user: dict = Depends(require_admin)):
    """Create a supervisor account (Admin only)"""
    user_id = str(uuid.uuid4())
    username = await get_next_username("supervisor", data.get("full_name", "موجه"))
    
    user_doc = {
        "id": user_id,
        "username": username,
        "password_hash": hash_password(data.get("password", "123456")),
        "full_name": data.get("full_name", ""),
        "role": "supervisor",
        "gender": data.get("gender", "male")
    }
    await db.users.insert_one(user_doc)
    
    return {"message": "تم إنشاء حساب الموجه", "username": username}

# ==================== GRADES ROUTES ====================

@api_router.get("/grades")
async def get_all_grades(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "student":
        student = await db.students.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if student:
            grades = await db.grades.find({"student_id": student["id"], "is_published": True}, {"_id": 0}).to_list(100)
            # For students, hide oral_skills and homework
            for grade_doc in grades:
                for g in grade_doc.get("grades", []):
                    g.pop("oral_skills", None)
                    g.pop("homework", None)
            return grades
        return []
    
    query = {}
    if current_user["role"] == "teacher":
        teacher = await db.teachers.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if teacher:
            assigned = teacher.get("assigned_classes", [])
            if assigned:
                query["$or"] = [{"class_name": a["class_name"], "section": a["section"]} for a in assigned]
    
    grades = await db.grades.find(query, {"_id": 0}).to_list(1000)
    return grades

@api_router.post("/grades")
async def create_or_update_grade(grade: GradeCreate, current_user: dict = Depends(require_teacher_or_above)):
    student = await db.students.find_one({"id": grade.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    total = grade.oral_skills + grade.homework + grade.quiz + grade.exam
    
    existing = await db.grades.find_one({
        "student_id": grade.student_id,
        "semester": grade.semester,
        "academic_year": grade.academic_year
    })
    
    grade_entry = {
        "subject": grade.subject,
        "oral_skills": grade.oral_skills,
        "homework": grade.homework,
        "quiz": grade.quiz,
        "exam": grade.exam,
        "total": total
    }
    
    if existing:
        grades_list = existing.get("grades", [])
        updated = False
        for i, g in enumerate(grades_list):
            if g["subject"] == grade.subject:
                grades_list[i] = grade_entry
                updated = True
                break
        if not updated:
            grades_list.append(grade_entry)
        
        await db.grades.update_one(
            {"student_id": grade.student_id, "semester": grade.semester, "academic_year": grade.academic_year},
            {"$set": {"grades": grades_list}}
        )
    else:
        grade_doc = {
            "id": str(uuid.uuid4()),
            "student_id": grade.student_id,
            "student_name": student["full_name"],
            "class_name": student["class_name"],
            "section": student["section"],
            "gender": student["gender"],
            "semester": grade.semester,
            "academic_year": grade.academic_year,
            "grades": [grade_entry],
            "is_published": False
        }
        await db.grades.insert_one(grade_doc)
    
    return {"message": "تم حفظ العلامات بنجاح"}

@api_router.post("/grades/publish")
async def publish_grades(data: dict, current_user: dict = Depends(require_admin_or_supervisor)):
    semester = data.get("semester")
    academic_year = data.get("academic_year")
    
    result = await db.grades.update_many(
        {"semester": semester, "academic_year": academic_year},
        {"$set": {"is_published": True}}
    )
    
    return {"message": f"تم نشر علامات {result.modified_count} طالب", "count": result.modified_count}

# ==================== SCHEDULES (الجداول) ====================

@api_router.get("/schedules/weekly")
async def get_weekly_schedule(class_name: Optional[str] = None, section: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if class_name:
        query["class_name"] = class_name
    if section:
        query["section"] = section
    
    schedules = await db.weekly_schedules.find(query, {"_id": 0}).to_list(1000)
    return schedules

@api_router.post("/schedules/weekly")
async def save_weekly_schedule(data: dict, current_user: dict = Depends(require_admin_or_supervisor)):
    schedule_doc = {
        "id": str(uuid.uuid4()),
        "class_name": data.get("class_name"),
        "section": data.get("section"),
        "items": data.get("items", []),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["full_name"]
    }
    
    # Upsert
    await db.weekly_schedules.update_one(
        {"class_name": data.get("class_name"), "section": data.get("section")},
        {"$set": schedule_doc},
        upsert=True
    )
    return {"message": "تم حفظ برنامج الأسبوع"}

@api_router.get("/schedules/exams")
async def get_exam_schedule(class_name: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if class_name:
        query["class_name"] = class_name
    
    schedules = await db.exam_schedules.find(query, {"_id": 0}).to_list(1000)
    return schedules

@api_router.post("/schedules/exams")
async def save_exam_schedule(data: dict, current_user: dict = Depends(require_admin_or_supervisor)):
    schedule_doc = {
        "id": str(uuid.uuid4()),
        "class_name": data.get("class_name"),
        "exam_type": data.get("exam_type", ""),  # midterm, final
        "items": data.get("items", []),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["full_name"]
    }
    
    await db.exam_schedules.update_one(
        {"class_name": data.get("class_name"), "exam_type": data.get("exam_type")},
        {"$set": schedule_doc},
        upsert=True
    )
    return {"message": "تم حفظ برنامج الامتحان"}

# ==================== CLASS FEES (أقساط الصفوف) ====================

@api_router.get("/class-fees")
async def get_class_fees(current_user: dict = Depends(require_admin_or_supervisor)):
    fees = await db.class_fees.find({}, {"_id": 0}).to_list(100)
    return fees

@api_router.post("/class-fees")
async def set_class_fee(data: ClassFee, current_user: dict = Depends(require_admin_or_supervisor)):
    await db.class_fees.update_one(
        {"class_name": data.class_name},
        {"$set": {"class_name": data.class_name, "annual_fee": data.annual_fee}},
        upsert=True
    )
    return {"message": f"تم تحديد قسط {data.class_name}"}

# ==================== ATTENDANCE ROUTES ====================

@api_router.get("/attendance")
async def get_attendance(date: Optional[str] = None, class_name: Optional[str] = None, gender: Optional[str] = None, current_user: dict = Depends(require_teacher_or_above)):
    query = {}
    if date:
        query["date"] = date
    if class_name:
        query["class_name"] = class_name
    if gender:
        # Need to join with students - filter after
        pass
    
    records = await db.attendance.find(query, {"_id": 0}).to_list(1000)
    
    if gender:
        # Filter by gender
        student_ids = [r["student_id"] for r in records]
        students = await db.students.find({"id": {"$in": student_ids}, "gender": gender}, {"_id": 0}).to_list(1000)
        valid_ids = {s["id"] for s in students}
        records = [r for r in records if r["student_id"] in valid_ids]
    
    return records

@api_router.post("/attendance")
async def mark_attendance(record: AttendanceCreate, current_user: dict = Depends(require_teacher_or_above)):
    student = await db.students.find_one({"id": record.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    existing = await db.attendance.find_one({
        "student_id": record.student_id,
        "date": record.date
    })
    
    if existing:
        await db.attendance.update_one(
            {"student_id": record.student_id, "date": record.date},
            {"$set": {"status": record.status, "notes": record.notes}}
        )
    else:
        await db.attendance.insert_one({
            "id": str(uuid.uuid4()),
            "student_id": record.student_id,
            "student_name": student["full_name"],
            "class_name": student["class_name"],
            "section": student["section"],
            "gender": student["gender"],
            "date": record.date,
            "status": record.status,
            "notes": record.notes
        })
    
    return {"message": "تم تسجيل الحضور"}

@api_router.post("/attendance/bulk")
async def mark_bulk_attendance(data: dict, current_user: dict = Depends(require_teacher_or_above)):
    records = data.get("records", [])
    date = data.get("date")
    
    for record in records:
        student = await db.students.find_one({"id": record["student_id"]}, {"_id": 0})
        if student:
            existing = await db.attendance.find_one({
                "student_id": record["student_id"],
                "date": date
            })
            
            if existing:
                await db.attendance.update_one(
                    {"student_id": record["student_id"], "date": date},
                    {"$set": {"status": record["status"]}}
                )
            else:
                await db.attendance.insert_one({
                    "id": str(uuid.uuid4()),
                    "student_id": record["student_id"],
                    "student_name": student["full_name"],
                    "class_name": student["class_name"],
                    "section": student["section"],
                    "gender": student["gender"],
                    "date": date,
                    "status": record["status"]
                })
    
    return {"message": f"تم تسجيل حضور {len(records)} طالب"}

# ==================== FINANCIAL ROUTES ====================

@api_router.get("/financials")
async def get_all_financials(search: Optional[str] = None, gender: Optional[str] = None, current_user: dict = Depends(require_admin_or_supervisor)):
    query = {}
    
    financials = await db.financials.find({}, {"_id": 0}).to_list(1000)
    
    # Get student info for gender filter
    if gender or search:
        result = []
        for f in financials:
            student = await db.students.find_one({"id": f["student_id"]}, {"_id": 0})
            if student:
                if gender and student.get("gender") != gender:
                    continue
                if search and search.lower() not in f.get("student_name", "").lower():
                    continue
            
            total_paid = sum(p.get("amount", 0) for p in f.get("payments", []))
            f["total_paid"] = total_paid
            f["remaining"] = f.get("total_fee", 0) - f.get("discount", 0) - total_paid
            f["gender"] = student.get("gender") if student else None
            result.append(f)
        return result
    
    for f in financials:
        total_paid = sum(p.get("amount", 0) for p in f.get("payments", []))
        f["total_paid"] = total_paid
        f["remaining"] = f.get("total_fee", 0) - f.get("discount", 0) - total_paid
        student = await db.students.find_one({"id": f["student_id"]}, {"_id": 0})
        f["gender"] = student.get("gender") if student else None
    
    return financials

@api_router.post("/financials")
async def create_or_update_financial(data: FinancialCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    student = await db.students.find_one({"id": data.student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="الطالب غير موجود")
    
    existing = await db.financials.find_one({"student_id": data.student_id})
    
    if existing:
        await db.financials.update_one(
            {"student_id": data.student_id},
            {"$set": {"total_fee": data.total_fee, "discount": data.discount}}
        )
    else:
        await db.financials.insert_one({
            "id": str(uuid.uuid4()),
            "student_id": data.student_id,
            "student_name": student["full_name"],
            "class_name": student["class_name"],
            "total_fee": data.total_fee,
            "discount": data.discount,
            "payments": []
        })
    
    return {"message": "تم حفظ البيانات المالية"}

@api_router.post("/financials/payment")
async def add_payment(data: PaymentCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    financial = await db.financials.find_one({"student_id": data.student_id})
    if not financial:
        raise HTTPException(status_code=404, detail="لا يوجد سجل مالي لهذا الطالب")
    
    receipt_number = f"RCP-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:4].upper()}"
    
    payment = {
        "id": str(uuid.uuid4()),
        "amount": data.amount,
        "date": data.date,
        "notes": data.notes,
        "receipt_number": receipt_number,
        "received_by": current_user["full_name"]
    }
    
    await db.financials.update_one(
        {"student_id": data.student_id},
        {"$push": {"payments": payment}}
    )
    
    return {"message": "تم تسجيل الدفعة", "receipt_number": receipt_number, "payment": payment}

@api_router.get("/financials/receipt/{student_id}/{payment_id}")
async def get_payment_receipt(student_id: str, payment_id: str, current_user: dict = Depends(require_admin_or_supervisor)):
    """Get receipt data for printing"""
    financial = await db.financials.find_one({"student_id": student_id}, {"_id": 0})
    if not financial:
        raise HTTPException(status_code=404, detail="لا يوجد سجل مالي")
    
    payment = None
    for p in financial.get("payments", []):
        if p.get("id") == payment_id:
            payment = p
            break
    
    if not payment:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")
    
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    
    return {
        "student_name": financial.get("student_name"),
        "class_name": financial.get("class_name"),
        "payment": payment,
        "parent_name": student.get("parent_info", {}).get("father_name") if student else ""
    }

# ==================== EXPORT ROUTES ====================

@api_router.get("/export/salaries")
async def export_salaries_excel(current_user: dict = Depends(require_admin_or_supervisor)):
    teachers = await db.teachers.find({}, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف الرواتب"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title with logo placeholder
    ws.merge_cells('A1:H1')
    ws['A1'] = "ثانوية ابن خلدون الخاصة - كشف رواتب الأساتذة"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['#', 'اسم الأستاذ', 'المادة', 'عدد الساعات', 'أجر الساعة', 'المكافآت', 'الخصومات', 'الراتب الصافي']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    total_salaries = 0
    for idx, teacher in enumerate(teachers, 1):
        salary = (teacher["hourly_rate"] * teacher["total_hours"]) + teacher["bonus"] - teacher["deductions"]
        total_salaries += salary
        
        row_data = [
            idx,
            teacher["full_name"],
            teacher["subject"],
            teacher["total_hours"],
            teacher["hourly_rate"],
            teacher["bonus"],
            teacher["deductions"],
            salary
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 4, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    total_row = len(teachers) + 5
    ws.merge_cells(f'A{total_row}:G{total_row}')
    ws[f'A{total_row}'] = "إجمالي الرواتب"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='center')
    ws[f'H{total_row}'] = total_salaries
    ws[f'H{total_row}'].font = Font(bold=True)
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15 if col != 'B' else 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=salaries_report.xlsx"}
    )

@api_router.get("/export/students-list")
async def export_students_list(class_name: str, section: str, gender: Optional[str] = None, current_user: dict = Depends(require_admin_or_supervisor)):
    """Export student list for a specific class/section"""
    query = {"class_name": class_name, "section": section}
    if gender:
        query["gender"] = gender
    
    students = await db.students.find(query, {"_id": 0}).sort("full_name", 1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة الطلاب"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:D1')
    ws['A1'] = "ثانوية ابن خلدون الخاصة"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:D2')
    gender_text = "ذكور" if gender == "male" else "إناث" if gender == "female" else "الكل"
    ws['A2'] = f"قائمة طلاب {class_name} - الشعبة {section} ({gender_text})"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['#', 'الاسم الكامل', 'اسم الأب', 'هاتف ولي الأمر']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for idx, student in enumerate(students, 1):
        row_data = [
            idx,
            student["full_name"],
            student.get("parent_info", {}).get("father_name", ""),
            student.get("parent_info", {}).get("father_phone", "")
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 4, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if col == 1 else 'right')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=students_{class_name}_{section}.xlsx"}
    )

@api_router.get("/export/grades-sheet")
async def export_grades_sheet(class_name: str, section: str, semester: str, academic_year: str, gender: Optional[str] = None, current_user: dict = Depends(require_admin_or_supervisor)):
    """Export grades sheet for a class/section"""
    query = {"class_name": class_name, "section": section, "semester": semester, "academic_year": academic_year}
    if gender:
        query["gender"] = gender
    
    grades = await db.grades.find(query, {"_id": 0}).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "كشف العلامات"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get all subjects
    all_subjects = set()
    for g in grades:
        for grade_entry in g.get("grades", []):
            all_subjects.add(grade_entry["subject"])
    subjects = sorted(list(all_subjects))
    
    ws.merge_cells(f'A1:{chr(65 + len(subjects) + 1)}1')
    ws['A1'] = "ثانوية ابن خلدون الخاصة - كشف العلامات"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    gender_text = "ذكور" if gender == "male" else "إناث" if gender == "female" else ""
    ws.merge_cells(f'A2:{chr(65 + len(subjects) + 1)}2')
    ws['A2'] = f"{class_name} - الشعبة {section} | {semester} - {academic_year} {gender_text}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['#', 'اسم الطالب'] + subjects
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for idx, grade_doc in enumerate(grades, 1):
        ws.cell(row=idx + 4, column=1, value=idx).border = border
        ws.cell(row=idx + 4, column=2, value=grade_doc["student_name"]).border = border
        
        grade_map = {g["subject"]: g["total"] for g in grade_doc.get("grades", [])}
        for col_idx, subj in enumerate(subjects, 3):
            cell = ws.cell(row=idx + 4, column=col_idx, value=grade_map.get(subj, ""))
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    for i, _ in enumerate(subjects):
        ws.column_dimensions[chr(67 + i)].width = 12
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=grades_{class_name}_{section}.xlsx"}
    )

@api_router.get("/export/financials")
async def export_financials_excel(gender: Optional[str] = None, current_user: dict = Depends(require_admin_or_supervisor)):
    financials = await db.financials.find({}, {"_id": 0}).to_list(1000)
    
    if gender:
        result = []
        for f in financials:
            student = await db.students.find_one({"id": f["student_id"]}, {"_id": 0})
            if student and student.get("gender") == gender:
                result.append(f)
        financials = result
    
    wb = Workbook()
    ws = wb.active
    ws.title = "الذمة المالية"
    ws.sheet_view.rightToLeft = True
    
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "ثانوية ابن خلدون الخاصة - كشف الذمة المالية"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    gender_text = "ذكور" if gender == "male" else "إناث" if gender == "female" else ""
    ws.merge_cells('A2:G2')
    ws['A2'] = f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d')} {gender_text}"
    ws['A2'].alignment = Alignment(horizontal='center')
    
    headers = ['#', 'اسم الطالب', 'الصف', 'القسط الكامل', 'الخصم', 'المدفوع', 'المتبقي']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    total_fees = 0
    total_paid = 0
    total_remaining = 0
    
    for idx, f in enumerate(financials, 1):
        paid = sum(p.get("amount", 0) for p in f.get("payments", []))
        remaining = f.get("total_fee", 0) - f.get("discount", 0) - paid
        
        total_fees += f.get("total_fee", 0)
        total_paid += paid
        total_remaining += remaining
        
        row_data = [
            idx,
            f.get("student_name", ""),
            f.get("class_name", ""),
            f.get("total_fee", 0),
            f.get("discount", 0),
            paid,
            remaining
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 4, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    total_row = len(financials) + 5
    ws.merge_cells(f'A{total_row}:C{total_row}')
    ws[f'A{total_row}'] = "الإجمالي"
    ws[f'A{total_row}'].font = Font(bold=True)
    ws[f'D{total_row}'] = total_fees
    ws[f'F{total_row}'] = total_paid
    ws[f'G{total_row}'] = total_remaining
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15 if col != 'B' else 25
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=financials_report.xlsx"}
    )

# ==================== ANNOUNCEMENTS ROUTES ====================

@api_router.get("/announcements")
async def get_announcements(current_user: dict = Depends(get_current_user)):
    query = {"is_active": True}
    
    # Filter by role
    if current_user["role"] == "student":
        query["target_audience"] = {"$in": ["all", "students"]}
    elif current_user["role"] == "teacher":
        query["target_audience"] = {"$in": ["all", "teachers"]}
    
    # Filter by gender if applicable
    user_gender = current_user.get("gender")
    if user_gender and current_user["role"] == "student":
        query["$or"] = [
            {"target_gender": None},
            {"target_gender": "all"},
            {"target_gender": user_gender}
        ]
    
    announcements = await db.announcements.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return announcements

@api_router.post("/announcements")
async def create_announcement(announcement: AnnouncementCreate, current_user: dict = Depends(require_admin_or_supervisor)):
    announcement_dict = announcement.model_dump()
    announcement_dict["id"] = str(uuid.uuid4())
    announcement_dict["created_at"] = datetime.now(timezone.utc).isoformat()
    announcement_dict["created_by"] = current_user["full_name"]
    announcement_dict["is_active"] = True
    
    await db.announcements.insert_one(announcement_dict)
    return announcement_dict

@api_router.delete("/announcements/{announcement_id}")
async def delete_announcement(announcement_id: str, current_user: dict = Depends(require_admin_or_supervisor)):
    result = await db.announcements.delete_one({"id": announcement_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الإعلان غير موجود")
    return {"message": "تم حذف الإعلان"}

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(require_admin_or_supervisor)):
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        settings = SchoolSettings().model_dump()
        await db.settings.insert_one(settings)
    return settings

@api_router.put("/settings")
async def update_settings(settings: SchoolSettings, current_user: dict = Depends(require_admin)):
    await db.settings.update_one({}, {"$set": settings.model_dump()}, upsert=True)
    return {"message": "تم تحديث الإعدادات"}

# ==================== DASHBOARD STATS ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(gender: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "student":
        student = await db.students.find_one({"user_id": current_user["id"]}, {"_id": 0})
        if student:
            grades = await db.grades.find({"student_id": student["id"], "is_published": True}, {"_id": 0}).to_list(100)
            attendance = await db.attendance.find({"student_id": student["id"]}, {"_id": 0}).to_list(1000)
            present_count = len([a for a in attendance if a["status"] == "present"])
            return {
                "grades_count": len(grades),
                "attendance_rate": round((present_count / len(attendance) * 100) if attendance else 0, 1)
            }
        return {}
    
    query = {}
    if gender:
        query["gender"] = gender
    
    students_count = await db.students.count_documents(query)
    teachers_count = await db.teachers.count_documents({})
    male_students = await db.students.count_documents({"gender": "male"})
    female_students = await db.students.count_documents({"gender": "female"})
    
    teachers = await db.teachers.find({}, {"_id": 0}).to_list(1000)
    total_salaries = sum(
        (t["hourly_rate"] * t["total_hours"]) + t["bonus"] - t["deductions"]
        for t in teachers
    )
    
    financials_query = {}
    if gender:
        student_ids = [s["id"] for s in await db.students.find({"gender": gender}, {"_id": 0, "id": 1}).to_list(1000)]
        financials_query["student_id"] = {"$in": student_ids}
    
    financials = await db.financials.find(financials_query, {"_id": 0}).to_list(1000)
    total_fees = sum(f.get("total_fee", 0) for f in financials)
    total_paid = sum(
        sum(p.get("amount", 0) for p in f.get("payments", []))
        for f in financials
    )
    total_remaining = total_fees - total_paid
    
    announcements_count = await db.announcements.count_documents({"is_active": True})
    
    return {
        "students_count": students_count,
        "teachers_count": teachers_count,
        "male_students": male_students,
        "female_students": female_students,
        "total_salaries": total_salaries,
        "total_fees": total_fees,
        "total_paid": total_paid,
        "total_remaining": total_remaining,
        "announcements_count": announcements_count
    }

# ==================== CLASSES DATA ====================

@api_router.get("/classes")
async def get_classes():
    return {
        "classes": [
            "الصف السابع",
            "الصف الثامن",
            "الصف التاسع",
            "الصف العاشر",
            "الصف الحادي عشر علمي",
            "الصف الحادي عشر أدبي",
            "الصف الثالث الثانوي علمي",
            "الصف الثالث الثانوي أدبي"
        ],
        "sections": ["الأولى", "الثانية", "الثالثة", "الرابعة"],
        "subjects": [
            "اللغة العربية",
            "اللغة الإنجليزية",
            "اللغة الفرنسية",
            "الرياضيات",
            "الفيزياء",
            "الكيمياء",
            "علم الأحياء",
            "التاريخ",
            "الجغرافيا",
            "الفلسفة",
            "التربية الوطنية",
            "التربية الدينية",
            "المعلوماتية"
        ],
        "semesters": ["الفصل الأول", "الفصل الثاني"],
        "academic_years": ["2024-2025", "2025-2026", "2026-2027"],
        "days": ["الأحد", "الاثنين", "الثلاثاء", "الأربعاء", "الخميس"],
        "periods": [1, 2, 3, 4, 5, 6, 7]
    }

# ==================== USERS LIST ====================

@api_router.get("/users/list")
async def get_users_list(current_user: dict = Depends(require_admin_or_supervisor)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_database():
    admin = await db.users.find_one({"username": "admin"})
    if admin:
        return {"message": "قاعدة البيانات مُعدة مسبقاً"}
    
    # Create Admin
    await db.users.insert_one({
        "id": str(uuid.uuid4()),
        "username": "admin",
        "password_hash": hash_password("admin123"),
        "full_name": "مدير المدرسة",
        "role": "admin",
        "gender": "male"
    })
    
    # Create Supervisor
    await db.users.insert_one({
        "id": str(uuid.uuid4()),
        "username": "موجه",
        "password_hash": hash_password("123456"),
        "full_name": "الموجه التربوي",
        "role": "supervisor",
        "gender": "male"
    })
    
    # Create default settings
    await db.settings.insert_one(SchoolSettings().model_dump())
    
    # Create default class fees
    default_fees = [
        {"class_name": "الصف السابع", "annual_fee": 400000},
        {"class_name": "الصف الثامن", "annual_fee": 400000},
        {"class_name": "الصف التاسع", "annual_fee": 450000},
        {"class_name": "الصف العاشر", "annual_fee": 500000},
        {"class_name": "الصف الحادي عشر علمي", "annual_fee": 550000},
        {"class_name": "الصف الحادي عشر أدبي", "annual_fee": 550000},
        {"class_name": "الصف الثالث الثانوي علمي", "annual_fee": 600000},
        {"class_name": "الصف الثالث الثانوي أدبي", "annual_fee": 600000},
    ]
    for fee in default_fees:
        await db.class_fees.insert_one(fee)
    
    return {
        "message": "تم إعداد قاعدة البيانات بنجاح",
        "credentials": {
            "admin": {"username": "admin", "password": "admin123"},
            "supervisor": {"username": "موجه", "password": "123456"}
        }
    }

# ==================== CLEAR DUMMY DATA ====================

@api_router.post("/clear-data")
async def clear_dummy_data(current_user: dict = Depends(require_admin)):
    """Clear all dummy/test data"""
    await db.students.delete_many({})
    await db.teachers.delete_many({})
    await db.grades.delete_many({})
    await db.attendance.delete_many({})
    await db.financials.delete_many({})
    await db.announcements.delete_many({})
    await db.weekly_schedules.delete_many({})
    await db.exam_schedules.delete_many({})
    await db.teacher_attendance.delete_many({})
    
    # Keep admin and supervisor users
    await db.users.delete_many({"role": {"$nin": ["admin", "supervisor"]}})
    
    return {"message": "تم مسح جميع البيانات التجريبية"}

# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "مرحباً بكم في نظام ثانوية ابن خلدون الخاصة"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
